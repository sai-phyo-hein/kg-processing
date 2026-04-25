"""Excel processing module for handling XLSX files."""

import os
import tempfile
from typing import List

try:
    from openpyxl import load_workbook
    from PIL import Image, ImageDraw, ImageFont
except ImportError as e:
    raise ImportError(f"Required dependencies not installed. Please install: {e}") from e


def _create_text_image(text: str, title: str = "") -> Image.Image:
    """Create an image from text content.

    Args:
        text: Text content to render
        title: Optional title for the image

    Returns:
        PIL Image with rendered text
    """
    # Create a white image
    width, height = 800, 600
    img = Image.new("RGB", (width, height), color="white")
    draw = ImageDraw.Draw(img)

    # Try to use a default font
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16)
        title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20)
    except (IOError, OSError):
        font = ImageFont.load_default()
        title_font = ImageFont.load_default()

    # Draw title if provided
    if title:
        draw.text((10, 10), title, fill="black", font=title_font)

    # Draw text content (simplified - in production, you'd want proper text wrapping)
    y_offset = 50
    for line in text.split("\n"):
        if y_offset < height - 30:
            draw.text((10, y_offset), line, fill="black", font=font)
            y_offset += 25

    return img


def process_xlsx(xlsx_path: str) -> List[str]:
    """Process Excel file and convert sheets to images.

    Args:
        xlsx_path: Path to the Excel file

    Returns:
        List of base64 encoded sheet images

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        PermissionError: If Excel file cannot be read
        IOError: If Excel file cannot be processed
    """
    import base64

    # Validate file exists
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    # Validate file is readable
    if not os.access(xlsx_path, os.R_OK):
        raise PermissionError(f"Cannot read Excel file: {xlsx_path}")

    base64_images = []

    try:
        # For Excel, we'll extract sheet content
        wb = load_workbook(xlsx_path)

        for sheet_num, sheet in enumerate(wb.worksheets):
            # Extract data from sheet
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    sheet_data.append(
                        "\t".join(str(cell) if cell is not None else "" for cell in row)
                    )

            # Create image from sheet content
            sheet_text = "\n".join(sheet_data) if sheet_data else f"Sheet {sheet_num + 1}"
            img = _create_text_image(sheet_text, f"Sheet {sheet_num + 1}")

            buffered = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            img.save(buffered, format="PNG")
            buffered.close()

            with open(buffered.name, "rb") as f:
                image_data = f.read()
                base64_images.append(base64.b64encode(image_data).decode("utf-8"))

            os.unlink(buffered.name)
    except Exception as e:
        raise IOError(f"Failed to process Excel file: {e}") from e

    return base64_images
